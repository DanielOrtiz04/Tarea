import openpyxl
import tkinter as tk
from tkinter import messagebox, simpledialog
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from tkinter import filedialog, simpledialog

#Modulo de Inventario
def listar_carro(text_widget):

    text_widget.delete('1.0', tk.END)

    book = openpyxl.load_workbook('vehiculos.xlsx')
    sheet = book['Inventario']

    headers = [cell.value for cell in sheet[1]]

    data_by_header = {header: [] for header in headers}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        for header, cell_value in zip(headers, row):
            data_by_header[header].append(cell_value)

    max_data_length = max(len(data) for data in data_by_header.values())

    for header in headers:
        text_widget.insert(tk.END, f"{header.ljust(12)}")
    text_widget.insert(tk.END, '\n')

    for i in range(max_data_length):
        for header, data in data_by_header.items():
            if i < len(data):
                text_widget.insert(tk.END, f"{str(data[i]).ljust(12)}")
            else:
                text_widget.insert(tk.END, "".ljust(12))
        text_widget.insert(tk.END, '\n')

def crear_carro(text_widget):
    book = openpyxl.load_workbook('.vehiculosxlsx')
    sheet = book['Inventario']

    CodigodelCarro = simpledialog.askinteger("Input", "Ingrese el código del Carro:")
    Marca = simpledialog.askstring("Input", "Ingrese Marca del carro:")
    Kilometraje = simpledialog.askinteger("Input", "Ingrese Km del Carro:")
    Modelo = simpledialog.askstring("Input", "Ingrese Modelo:")
    precio = simpledialog.askstring("Input", "Ingrese el precio del carro:")

    data = [CodigodelCarro, Marca, Kilometraje, Modelo,precio]
    sheet.append(data)

    book.save('vehiculos.xlsx')

def actualizar_carro(text_widget):
    try:
        book = openpyxl.load_workbook('vehiculos.xlsx')
        sheet = book['Inventario']

        codigo = simpledialog.askinteger("Input", "Ingrese el código del carro que desea actualizar:")

        producto_encontrado = False

        for fila in sheet.iter_rows(min_col=1, max_col=1):
            celda = fila[0]
            if celda.value == codigo:
                text_widget.insert(tk.END, "Se ha encontrado el producto\n")
                producto_encontrado = True
                fila_destino = celda.row 
                columna_destino = 5
                celda_destino = sheet.cell(row=fila_destino, column=columna_destino)
                precio_nuevo = simpledialog.askstring("Input", "Ingrese el nuevo precio:")
                celda_destino.value = precio_nuevo
                break 

        if not producto_encontrado:
            text_widget.insert(tk.END, "No se encontró el producto con el código proporcionado\n")

        book.save('vehiculos.xlsx')

    except Exception as e:
        text_widget.insert(tk.END, f"Se produjo un error: {str(e)}\n")

def eliminar_carro(text_widget):
    try:
        book = openpyxl.load_workbook('vehiculos.xlsx')
        sheet = book['Inventario']

        codigo = simpledialog.askinteger("Input", "Ingrese el código del carro que desea eliminar:")

        producto_encontrado = False

        for fila in sheet.iter_rows(min_col=1, max_col=1):
            celda = fila[0]
            if celda.value == codigo:
                text_widget.insert(tk.END, "Se ha encontrado el producto\n")
                producto_encontrado = True
                fila_destino = celda.row 
                sheet.delete_rows(fila_destino)
                text_widget.insert(tk.END, "El producto ha sido eliminado.\n")
                break 

        if not producto_encontrado:
            text_widget.insert(tk.END, "No se encontró el carro con el código proporcionado\n")

        book.save('vehiculos.xlsx')

    except Exception as e:
        text_widget.insert(tk.END, f"Se produjo un error: {str(e)}\n")

#Modulo Clientes

def listar_carro(text_widget):

    text_widget.delete('1.0', tk.END)

    book = openpyxl.load_workbook('vehiculos.xlsx')
    sheet = book['Clientes']

    headers = [cell.value for cell in sheet[1]]

    data_by_header = {header: [] for header in headers}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        for header, cell_value in zip(headers, row):
            data_by_header[header].append(cell_value)

    max_data_length = max(len(data) for data in data_by_header.values())

    for header in headers:
        text_widget.insert(tk.END, f"{header.ljust(12)}")
    text_widget.insert(tk.END, '\n')

    for i in range(max_data_length):
        for header, data in data_by_header.items():
            if i < len(data):
                text_widget.insert(tk.END, f"{str(data[i]).ljust(12)}")
            else:
                text_widget.insert(tk.END, "".ljust(12))
        text_widget.insert(tk.END, '\n')


def crear_carro(text_widget):
    book = openpyxl.load_workbook('vehiculos.xlsx')
    sheet = book['Clientes']

    while True:
        codigo = simpledialog.askinteger("Input", "Ingrese el código del carro a agregar: ")
        
        if codigo == -1:
            break

        nombre = simpledialog.askstring("Input", "Ingrese el nombre del carro:")

        clientes_existentes = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            codigo_cliente, nombre_cliente, direccion_cliente = row
            clientes_existentes.append((codigo_cliente, nombre_cliente, direccion_cliente))

        if any(cliente[0] == codigo for cliente in clientes_existentes):
            text_widget.insert(tk.END, f"El cliente con código {codigo} ya existe. Introduzca otro código.\n")
        else:
            nuevo_cliente = (codigo, nombre, Direccion)
            clientes_existentes.append(nuevo_cliente)

            clientes_ordenados = sorted(clientes_existentes, key=lambda cliente: cliente[0])

            for _ in range(2, sheet.max_row + 1):
                sheet.delete_rows(2)

            for cliente in clientes_ordenados:
                sheet.append(cliente)

            text_widget.insert(tk.END, "El nuevo cliente fue agregado de forma correcta\n")

            book.save('vehiculos.xlsx')
            break

def actualizar_carro(text_widget):
    try:
        book = openpyxl.load_workbook('vehiculos.xlsx')
        sheet = book['Clientes']

        codigo = simpledialog.askinteger("Input", "Ingrese el código del carro que desea actualizar:")

        cliente_encontrado = False

        for fila in sheet.iter_rows(min_col=1, max_col=1):
            celda = fila[0]
            if celda.value == codigo:
                text_widget.insert(tk.END, "Se ha encontrado el cliente\n")
                cliente_encontrado = True
                fila_destino = celda.row 
                columna_destino = 3
                celda_destino = sheet.cell(row=fila_destino, column=columna_destino)
                celda_destino.value = direccion_nueva
                text_widget.insert(tk.END, "Los datos del carro se actualizaron correctamente\n")
                break 

        if not cliente_encontrado:
            text_widget.insert(tk.END, "No se encontró el carro con el código proporcionado\n")

        book.save('vehiculos.xlsx')

    except Exception as e:
        text_widget.insert(tk.END, f"Se produjo un error: {str(e)}\n")

def eliminar_carro(text_widget):
    try:
        book = openpyxl.load_workbook('inventario.xlsx')
        sheet = book['Clientes']

        codigo = simpledialog.askinteger("Input", "Ingrese el código del carro que desea eliminar:")

        cliente_encontrado = False

        for fila in sheet.iter_rows(min_col=1, max_col=1):
            celda = fila[0]
            if celda.value == codigo:
                text_widget.insert(tk.END, "Se ha encontrado el carro\n")
                cliente_encontrado = True
                fila_destino = celda.row 
                sheet.delete_rows(fila_destino) 
                text_widget.insert(tk.END, "El carro ha sido eliminado.\n")
                break 

        if not cliente_encontrado:
            text_widget.insert(tk.END, "No se encontró el carro con el código proporcionado\n")

        book.save('vehiculos.xlsx')

    except Exception as e:
        text_widget.insert(tk.END, f"Se produjo un error: {str(e)}\n")


#Modulo de Ventas

def listar_carro(text_widget):

    text_widget.delete('1.0', tk.END)

    book = openpyxl.load_workbook('vehiculos.xlsx')
    sheet = book['Ventas']

    headers = [cell.value for cell in sheet[1]]

    data_by_header = {header: [] for header in headers}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        for header, cell_value in zip(headers, row):
            data_by_header[header].append(cell_value)

    max_data_length = max(len(data) for data in data_by_header.values())

    for header in headers:
        text_widget.insert(tk.END, f"{header.ljust(14)}")
    text_widget.insert(tk.END, '\n')

    for i in range(max_data_length):
        for header, data in data_by_header.items():
            if i < len(data):
                text_widget.insert(tk.END, f"{str(data[i]).ljust(20)}")
            else:
                text_widget.insert(tk.END, "".ljust(12))
        text_widget.insert(tk.END, '\n')

#Modulo gmail

def enviar_correo(text_widget):
    servidor_smtp = 'smtp.gmail.com'  
    puerto = 587  
    usuario = 'miusuario'
    contrasena = 'micontraseña'
    destinatario = simpledialog.askstring("Input", "Ingrese el correo electrónico del destinatario:")
    asunto = 'Informe de ventas'
    cuerpo = 'Aquí está el informe de ventas que solicitaste.'
    archivo = filedialog.askopenfilename(title="Seleccione el archivo a enviar")

    msg = MIMEMultipart()
    msg['From'] = usuario
    msg['To'] = destinatario
    msg['Subject'] = asunto

    msg.attach(MIMEText(cuerpo, 'plain'))

    adjunto = open(archivo, 'rb')

    part = MIMEBase('application', 'octet-stream')
    part.set_payload((adjunto).read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', 'attachment; filename= ' + archivo)

    msg.attach(part)

    servidor = smtplib.SMTP(servidor_smtp, puerto)
    servidor.starttls()
    servidor.login(usuario, contrasena)
    text = msg.as_string()
    servidor.sendmail(usuario, destinatario, text)
    servidor.quit()


def main():
    root = tk.Tk()
    root.title("Inventario")

    frame1 = tk.Frame(root)
    frame1.pack()

    frame2 = tk.Frame(root)
    frame2.pack()

    text_widget = tk.Text(frame2)
    text_widget.pack()

    # Crea un menú desplegable
    menu = tk.Menu(root)
    root.config(menu=menu)

    # Crea el menú de inventario y añádelo al menú desplegable
    inventario_menu = tk.Menu(menu)
    menu.add_cascade(label="Inventario", menu=inventario_menu)

    #Botones menú inventario
    inventario_menu.add_command(label="Listar Carro", command=lambda: listar_producto(text_widget))
    inventario_menu.add_command(label="Crear Carro", command=lambda: crear_carro(text_widget))
    inventario_menu.add_command(label="Actualizar Carro", command=lambda: actualizar_carro(text_widget))
    inventario_menu.add_command(label="Actualizar Carro", command=lambda: actualizar_existencia(text_widget))
    inventario_menu.add_command(label="Eliminar Carro", command=lambda: eliminar_producto(text_widget))

    #Menú clientes
    clientes_menu = tk.Menu(menu)
    menu.add_cascade(label="Carros", menu=clientes_menu)

    #Botones menú clientes
    clientes_menu.add_command(label="Listar Carr0", command=lambda: listar_clientes(text_widget))
    clientes_menu.add_command(label="Crear Carro", command=lambda: crear_clientes(text_widget))
    clientes_menu.add_command(label="Actualizar Carro", command=lambda: actualizar_cliente(text_widget))
    clientes_menu.add_command(label="Eliminar Carro", command=lambda: eliminar_cliente(text_widget))

    #Menu Ventas
    ventas_menu = tk.Menu(menu)
    menu.add_cascade(label="Listar", menu=ventas_menu)

    #Botones menú ventas
    ventas_menu.add_command(label="Listar Carro", command=lambda: listar_ventas(text_widget))
    ventas_menu.add_command(label="Agregar Carro", command=lambda: agregar_venta(text_widget))

    root.mainloop()

if __name__ == "__main__":
    main()
