import argparse
import openpyxl
import sys

#Modulo de Inventario

def listar_carro():
    book = openpyxl.load_workbook('vehiculos.xlsx')
    sheet = book['Inventario']

    headers = [cell.value for cell in sheet[1]]

    data_by_header = {header: [] for header in headers}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        for header, cell_value in zip(headers, row):
            data_by_header[header].append(cell_value)

    max_data_length = max(len(data) for data in data_by_header.values())

    for header in headers:
        print(f"{header.ljust(12)}", end='')
    print()  

    for i in range(max_data_length):
        for header, data in data_by_header.items():
            if i < len(data):
                print(f"{str(data[i]).ljust(12)}", end='')
            else:
                print("".ljust(12), end='')  
        print() 

def crear_carro():
    book = openpyxl.load_workbook('vehiculos.xlsx')
    sheet = book['Inventario']

    CodigodelCarro = int(input("Ingrese el código del Carro: "))
    Marca = input("Ingrese Marca del Carro: ")
    Kilometraje = int(input("Ingrese Km del carro: "))
    Modelo = input("Ingrese Modelo: ")
    precio = input("Ingrese el precio del carro: ")

    data = [CodigodelCarro, Marca, Kilometraje, Modelo, precio]
    sheet.append(data)

    book.save('inventario.xlsx')

def eliminar_carro():
    try:
        book = openpyxl.load_workbook('vehiculos.xlsx')
        sheet = book['Inventario']

        codigo = int(input("Ingrese el Codigo del Carro: "))

        producto_encontrado = False

        for fila in sheet.iter_rows(min_col=1, max_col=1):
            celda = fila[0]
            if celda.value == codigo:
                print("Se ha encontrado el producto")
                producto_encontrado = True
                fila_destino = celda.row 
                sheet.delete_rows(fila_destino)
                print("El producto ha sido eliminado.")
                break 

        if not producto_encontrado:
            print("No se encontró el producto con el código proporcionado")

        book.save('vehiculos.xlsx')

    except Exception as e:
        print("Se produjo un error:", str(e))



#Modulo de clientes

def listar_carro():
    book = openpyxl.load_workbook('vehiculos.xlsx')
    sheet = book['Clientes']

    headers = [cell.value for cell in sheet[1]]

    data_by_header = {header: [] for header in headers}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        for header, cell_value in zip(headers, row):
            data_by_header[header].append(cell_value)

    max_data_length = max(len(data) for data in data_by_header.values())

    for header in headers:
        print(f"{header.ljust(12)}", end='')
    print()  

    for i in range(max_data_length):
        for header, data in data_by_header.items():
            if i < len(data):
                print(f"{str(data[i]).ljust(12)}", end='')
            else:
                print("".ljust(12), end='')  
        print()

def crear_carro():
    book = openpyxl.load_workbook('vehiculos.xlsx')
    sheet = book['Clientes']

    while True:
        codigo = int(input("Ingrese el código del carro (o -1 para salir): "))
        
        if codigo == -1:
            break

        nombre = input("Ingrese el nombre del carro: ")

        clientes_existentes = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            codigo_carro, nombre, = row
            clientes_existentes.append((codigo_carro, nombre_carro,))

        if any(cliente[0] == codigo for cliente in clientes_existentes):
            print(f"El carro con código {codigo} ya existe. Introduzca otro código.")
        else:
            nuevo_carro = (codigo, Marca, )
            clientes_existentes.append(nuevo_cliente)

            clientes_ordenados = sorted(clientes_existentes, key=lambda cliente: cliente[0])

            for _ in range(2, sheet.max_row + 1):
                sheet.delete_rows(2)

            for cliente in clientes_ordenados:
                sheet.append(cliente)

            print("El nuevo carro fue agregado de forma correcta")

            book.save('vehiculos.xlsx')
            break



def eliminar_carro():
    try:
        book = openpyxl.load_workbook('vehiculos.xlsx')
        sheet = book['Clientes']
        codigo = int(input("Ingrese el código del carro que desee Eliminar: "))

        cliente_encontrado = False

        for fila in sheet.iter_rows(min_col=1, max_col=1):
            celda = fila[0]
            if celda.value == codigo:
                print("Se ha encontrado el carro")
                cliente_encontrado = True
                fila_destino = celda.row 
                sheet.delete_rows(fila_destino) 
                print("El carro ha sido eliminado.")
                break 

        if not cliente_encontrado:
            print("No se encontró el carro con el código proporcionado")

        book.save('vehiculos.xlsx')

    except Exception as e:
        print("Se produjo un error:", str(e))

#Modulo de ventas

def verificar_producto_en_inventario(codigo_producto, hoja_inventario):
    for row in hoja_inventario.iter_rows(values_only=True):
        if row[0] == codigo_producto:
            return row

    return None

def agregar_venta():
    codigo_producto = int(input("Ingrese el código del Carro a vender: "))

    book = openpyxl.load_workbook('vehiculos.xlsx')
    
    hoja_inventario = book['Inventario']

    producto = verificar_producto_en_inventario(codigo_producto, hoja_inventario)

    if producto is None:
        print(f"El producto con código {codigo_producto} no se encuentra en el inventario.")
        return

    cantidad_existente = producto[2] 
    precio_unitario = producto[4]

    if cantidad_existente <= 0:
        print("El producto está agotado y no se puede vender.")
        return

    cantidad_vendida = int(input(f"Ingrese la cantidad a vender (existencia actual: {cantidad_existente}): "))

    if cantidad_vendida > cantidad_existente:
        print("No hay suficiente cantidad en inventario para la venta.")
        return

    total_venta = cantidad_vendida * precio_unitario

    hoja_ventas = book['Ventas']
    hoja_ventas.append([codigo_producto, codigo_cliente, cantidad_vendida, total_venta])

    for idx, row in enumerate(hoja_inventario.iter_rows(values_only=True, min_row=2), start=2):
        if row[0] == codigo_producto:
            cantidad_actual = row[2]
            nueva_cantidad = cantidad_actual - cantidad_vendida
            hoja_inventario.cell(row=idx, column=3, value=nueva_cantidad)

    book.save('vehiculos.xlsx')

    print(f"Venta registrada exitosamente. Total de la venta: {total_venta}.")



#Modulo de consultas

 
    
def menu_interactivo():
    while True:
        print("Menú de opciones:")
        print("1. Listar Carro")
        print("2. Crear Carro")
        print("3. Actualizar Carro")
        print("4. Eliminar Carro")
        print("5. Listar carro")
        print("6. Crear carro")
        print("7. Actualizar carrro")
        print("8. Eliminar carro")
        print("9. Listar carro")
        print("10. Agregar venta") 
        print("11. Salir")
        
        opcion = input("Ingrese el número de la opción que desee: ")
        
        if opcion == '1':
            listar_producto()
        elif opcion == '2':
            crear_producto()
        elif opcion == '3':
            actualizar_producto()
        elif opcion == '4':
            actualizar_existencia()
        elif opcion == '5':
            eliminar_producto()
        elif opcion == '6':
            listar_clientes()
        elif opcion == '7':
            crear_clientes()
        elif opcion == '8':
            actualizar_cliente()
        elif opcion == '9':
            eliminar_cliente()
        elif opcion == '10':
            listar_ventas()
        elif opcion == '11':
            agregar_venta()
            break
        else:
            print("Opción no válida. Intente nuevamente.")

        continuar = input("¿Desea ejecutar otra función? (S/N): ")
        if continuar.lower() != 's':
            break

def linea_de_comandos():
    parser = argparse.ArgumentParser()
    parser.add_argument('--inventario_listar', action='store_true')
    parser.add_argument('--ayuda', action='store_true')
    parser.add_argument('--inventario_crear', action='store_true')
    parser.add_argument('--inventario_actualizar', action='store_true')
    parser.add_argument('--inventario_existencia', action='store_true')
    parser.add_argument('--inventario_eliminar', action='store_true')
    parser.add_argument('--listar_clientes', action='store_true')
    parser.add_argument('--crear_clientes', action='store_true')
    parser.add_argument('--actualizar_cliente', action='store_true')
    parser.add_argument('--eliminar_clientes', action='store_true')
    parser.add_argument('--listar_ventas', action='store_true')
    parser.add_argument('--agregar_venta', action='store_true')
    parser.add_argument('--anular_ventas', action='store_true')
    args = parser.parse_args()

    if args.ayuda:
        Ayuda()
    elif args.inventario_listar:
        listar_producto()
    elif args.inventario_crear:
        crear_producto()
    elif args.inventario_actualizar:
        actualizar_producto()
    elif args.inventario_existencia:
        actualizar_existencia()
    elif args.inventario_eliminar:
        eliminar_producto()
    elif args.listar_clientes:
        listar_clientes()
    elif args.crear_clientes:
        crear_clientes()
    elif args.actualizar_cliente:
        actualizar_cliente()
    elif args.eliminar_clientes:
        eliminar_cliente()
    elif args.listar_ventas:
        listar_ventas()
    elif args.agregar_venta:
        agregar_venta()
    elif args.anular_ventas:
        anular_venta()

if __name__ == "__main__":
    if len(sys.argv) == 1:
        menu_interactivo()
    else:
        linea_de_comandos()